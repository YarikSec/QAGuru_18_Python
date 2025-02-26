from zipfile import ZipFile
import pytest
import csv
from pypdf import PdfReader
from xlrd import open_workbook # Для xls
from openpyxl import load_workbook # Для xlsx

from homework_7.tests.conftest import ZIP_PATH



def test_read_csv_file_zip():
    with ZipFile(ZIP_PATH) as zip_file:
        csv_content = zip_file.read("file_example_XLS.csv").decode("utf-8-sig")
        reader = csv.reader(csv_content.splitlines(), delimiter=";")
        data = list(reader)
        assert "First Name" in data[0][1]
        assert "Last Name" in data[0][2]

def test_read_xls_file_zip():
    with ZipFile(ZIP_PATH) as zip_file:
        xls_content = zip_file.read("file_example_XLS_10.xls")
        xls_workbook = open_workbook(file_contents=xls_content)
        sheet = xls_workbook.sheet_by_index(0)
        assert sheet.nrows > 0
        assert sheet.ncols > 0



def test_read_xlsx_file_zip():
    with ZipFile(ZIP_PATH) as zip_file:
        with zip_file.open("file_example_XLSX_50.xlsx") as xlsx:
            xlsx_workbook = load_workbook(xlsx)
            sheet = xlsx_workbook.active
            assert sheet.max_row > 0
            assert sheet.max_column > 0

def test_read_pdf_file_zip():
    with ZipFile(ZIP_PATH) as zip_file:
        with zip_file.open("Python Testing with Pytest (Brian Okken).pdf") as pdf:
            reader = PdfReader(pdf)
            assert "Simple, Rapid, Effective, and Scalable" in reader.pages[1].extract_text()
            assert "Python Testing with pytest" in reader.pages[1].extract_text()